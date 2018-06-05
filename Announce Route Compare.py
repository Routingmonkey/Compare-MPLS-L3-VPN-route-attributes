import textfsm
import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import *
import tkFileDialog
from tkFileDialog import askdirectory
import tkMessageBox
import os
import os.path

#this function is used to just create simple list 
##e.g [PE name , Network , subnet mask , vrf , SOO , LP ,RT,RT]
#this function is called in function createtablefromtext 
def writedataincvsfile ( fsm_result , name , role ): 
 k = []
 for textrow in fsm_result:

  for i in textrow[-1]:
   j = i.split("RT:")
   for m in j:
    if m != "":
      k.append(m.strip())
  textrow.pop()
  for x in k:
   textrow.append(x)
  k = []

#this function is to used to capture required data from the text file and returns list of list 
#e.g [PE name , Network , subnet mask , vrf , SOO , LP ,[RT,RT]]
#this function is called in the function compareroutes
def createtablefromtext ( peroutetext , role , flag):
 PeRoutesFile = open( peroutetext , "r")
 raw_text_data = PeRoutesFile.read()
 re_table = textfsm.TextFSM(template)
 fsm_result = re_table.ParseText(raw_text_data)
 if flag == "KO":
	fsm_result.pop()
 writedataincvsfile ( fsm_result , "Routes Tagging" , role)
 return fsm_result


def compareroutes():

	global CENAME
	global precheckfilepath
	global postcheckfilepath
	global outputlocationpath
	global template
	
	# Open Excel sheet which is already prepared with conditional formatting to highlight unique values by comparing
    # ODD (pre check) and EVEN (post check) rows	
	wb = openpyxl.load_workbook('RT compare.xlsx',keep_vba=True, keep_links=True)
	name = "Routes Tagging"
	sheet = wb.get_sheet_by_name(name)
	template = open("show_route.textfsm") # This is textfsm regular expression template to capture SOO,RT,Local Perf for test file
	
	CENAME = e1.get().strip()
	if CENAME == "":
		tkMessageBox.showerror("Error", "Missing CE name")
		return

	precheckfilepath = precheckfilepathtext['text']  
	if precheckfilepathtext['text'] == "Select PRECHCK file" or precheckfilepathtext['text'] == "" :
		tkMessageBox.showerror("Error", "PRECHECK file path not provided")
		return

	postcheckfilepath = postcheckfilepathtext['text']  
	if postcheckfilepathtext['text'] == "Select POSTCHECK file" or postcheckfilepathtext['text'] == "" :
		tkMessageBox.showerror("Error", "POSTCHECK file path not provided")
		return
		
		
	outputlocationpath = outputlocationpathtext['text']  
	if outputlocationpathtext['text'] == "Save output to" or outputlocationpathtext['text'] == "" :
		tkMessageBox.showerror("Error", "Output path not provided")
		return
	
	path = outputlocationpath + "/" + CENAME + "_Announce Route compmare.xlsm"
	
	if os.path.isfile(path):
		tkMessageBox.showerror("Error", "Output file already exits")
		return
	
	temppeold = createtablefromtext ( precheckfilepath , "MAIN" , "KO") # this function call returns PRECHECK data in python list format
	temppenew = createtablefromtext ( postcheckfilepath , "BACKUP" , "KO") # this function call returns POSTCHECK data in python list format

	irow = 9  # used for precheck 
	jrow = 10 # used for postcheck
    
	# below for loops are just to match network,subnet mask,vrf from pre and post check python list that we got from function createtablefromtext
	# and write data for precheck in ODD row starting from 9 
	# and write data for precheck in EVEN row starting from 10
	for i in temppeold:
		for j in temppenew:
			x = 0
			if i[2] == j[2] and i[1] == j[1] and i[3] == j[3]:
				excelcol = 1
				for z in i: 
					sheet[get_column_letter(excelcol) + str(irow)] = z
					excelcol = excelcol + 1
				irow = irow + 2
				excelcol = 1
				for z in j:
					sheet[get_column_letter(excelcol) + str(jrow)] = z
					excelcol = excelcol + 1
				jrow = jrow + 2			
				x = x + 1
				excelcol = 1

	wb.save(path)


master = Tk()

master.iconbitmap('Monkey.ico')  #ICON for title bar

master.title('Announce Route Compare') # Title for GUI window
master.geometry('900x350') # size of GUI window

def getprecheckfiletpath():	
	precheckpath = tkFileDialog.askopenfilename()
	precheckfilepathtext.config(text=precheckpath)
	
def getpostcheckfilepath():
	postcheckpath = tkFileDialog.askopenfilename()
	postcheckfilepathtext.config(text=postcheckpath)

def getoutputlocationpath():	
	scriptoutputpath = tkFileDialog.askdirectory()
	outputlocationpathtext.config(text=scriptoutputpath)

Label(master, text="CE Name :", font = ('Comic Sans MS',12)).grid(row=0, column=12)  # Label for CE name entry
Label(master, text="PRECHCK file path :", font = ('Comic Sans MS',12)).grid(row=1, column=12) # Label for precheck file path button  
Label(master, text="POETCHECK file path :", font = ('Comic Sans MS',12)).grid(row=2, column=12) # Label for postcheck file path button 
Label(master, text="OUTPUT path :", font = ('Comic Sans MS',12)).grid(row=3, column=12) # Label for output file path button 

e1 = Entry(master)
e1.grid(row=0, column=15) # Entry option to manually type the CE name

#Button to call getprecheckfiletpath function to get precheck file path
precheckfilepathtext = Button(master, text='Select PRECHCK file', command=getprecheckfiletpath, font = ('Comic Sans MS',8)) 
precheckfilepathtext.grid(row=1, column=15)

#Button to call getpostcheckfilepath function to get postcheck file path
postcheckfilepathtext = Button(master, text='Select POSTCHECK file', command=getpostcheckfilepath, font = ('Comic Sans MS',8))
postcheckfilepathtext.grid(row=2, column=15)

#Button to call getoutputlocationpath function to provide output file location
outputlocationpathtext = Button(master, text='Save output to', command=getoutputlocationpath, font = ('Comic Sans MS',8))
outputlocationpathtext.grid(row=3, column=15)

#Quit button to quit the program
Button(master, text='Quit', command=master.quit, font = ('Comic Sans MS',12)).grid(row=20, column=5, sticky=W, pady=4)

#Compare button to call compareroutes function
Button(master, text='Compare', command=compareroutes, font = ('Comic Sans MS',12)).grid(row=20, column=20, sticky=W, pady=4) 

mainloop( )